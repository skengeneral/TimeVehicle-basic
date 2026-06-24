import webbrowser
import sys
import os
import time
import requests
import subprocess

from pathlib import Path

from playwright.sync_api import sync_playwright

from PyQt6.QtWidgets import (
    QApplication, QWidget, QLabel, QLineEdit, QCheckBox,
    QPushButton, QVBoxLayout, QHBoxLayout, QFrame, QMessageBox,
    QScrollArea, QInputDialog, QTextEdit
)
from PyQt6.QtCore import Qt, QPoint, QTimer, QThread, pyqtSignal
from PyQt6.QtGui import QFont, QPainter, QPolygon, QColor, QCursor, QIcon


# ── API Key loader ────────────────────────────────────────────────
def get_local_api_key():
    if getattr(sys, 'frozen', False):
        current_dir = os.path.dirname(sys.executable)
    else:
        current_dir = os.path.dirname(os.path.abspath(__file__))
    key_file_path = os.path.join(current_dir, "serp_api.txt")
    if os.path.exists(key_file_path):
        try:
            with open(key_file_path, "r", encoding="utf-8") as file:
                return file.read().strip()
        except: pass
    return os.environ.get("SERPAPI_KEY")

def fetch_live_serp_credits():
    api_key = get_local_api_key()
    if not api_key:
        return "Missing Key"
    endpoint = "https://serpapi.com/account.json"
    params = {"api_key": api_key}
    try:
        response = requests.get(endpoint, params=params, timeout=3)
        if response.status_code == 200:
            account_info = response.json()
            return str(account_info.get("plan_searches_left", 0))
    except:
        pass
    return "Offline (Check Connection)"


# ── Passkey cache helpers ─────────────────────────────────────────
# Clients enter their passkey once. It is validated against the cloud
# sheet and saved locally. Every subsequent search silently re-checks
# the saved key — no prompt unless the key has been revoked.

def _auth_file_path():
    """Returns path to the local passkey cache file."""
    base = (
        os.path.dirname(sys.executable)
        if getattr(sys, 'frozen', False)
        else os.path.dirname(os.path.abspath(__file__))
    )
    return os.path.join(base, ".tvauth")

def _load_saved_passkey():
    try:
        with open(_auth_file_path(), 'r', encoding='utf-8') as f:
            return f.read().strip()
    except Exception:
        return None

def _save_passkey(passkey):
    try:
        with open(_auth_file_path(), 'w', encoding='utf-8') as f:
            f.write(passkey)
    except Exception:
        pass

def _clear_saved_passkey():
    try:
        os.remove(_auth_file_path())
    except Exception:
        pass

def _validate_passkey_cloud(passkey_clean):
    """
    Checks passkey against Google Sheet.
    Returns: True  = valid key
             False = invalid / revoked key
             None  = offline / server unreachable (can't determine)
    """
    CSV_URL = (
        "https://docs.google.com/spreadsheets/d/"
        "1_mHFrZcnhupYNU2FA9B1I5DEFNerNsIwW61lX_ygPHs/export?format=csv&gid=0"
    )
    try:
        resp = requests.get(CSV_URL, timeout=4)
        if resp.status_code == 200:
            active_keys = set()
            for line in resp.text.splitlines():
                for segment in line.split(','):
                    token = segment.strip().replace('"', '').replace("'", "")
                    if token:
                        active_keys.add(token.lower())
            return passkey_clean.lower() in active_keys
        return None   # server error → treat as offline
    except Exception:
        return None   # network error → offline


# ── Logo widget ───────────────────────────────────────────────────
class LogoWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedSize(50, 50)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        painter.setBrush(QColor("#6B1D66"))
        painter.setPen(Qt.PenStyle.NoPen)
        painter.drawRect(0, 0, self.width(), self.height())
        painter.setBrush(QColor("#E397E1"))
        points = QPolygon([
            QPoint(0, 0), QPoint(self.width(), 0),
            QPoint(self.width() // 2, int(self.height() * 0.46))
        ])
        painter.drawPolygon(points)


# ── Background scraper thread ─────────────────────────────────────
class ScraperWorker(QThread):
    """
    Fetches the latest scraper_engine.py from GitHub and executes it
    in RAM so the client always runs the most up-to-date logic without
    any local updates.
    """
    finished_signal = pyqtSignal(dict)
    error_signal    = pyqtSignal(str)
    progress_signal = pyqtSignal(str)   # ← NEW: live log messages

    def __init__(self, search_query, allowed_ratings, target_city):
        super().__init__()
        self.search_query    = search_query
        self.allowed_ratings = allowed_ratings
        self.target_city     = target_city

    def run(self):
        try:
            headers = {'User-Agent': 'Mozilla/5.0'}
            URL = (
                f"https://raw.githubusercontent.com/skengeneral/TimeVehicle-basic"
                f"/main/scraper_engine.py?t={int(time.time())}"
            )
            self.progress_signal.emit("🌐 Fetching latest engine from cloud...")
            response = requests.get(URL, headers=headers, timeout=15)

            if response.status_code == 200:
                raw_code = response.text

                # Patch __file__ references for frozen / exec context
                patched_code = "from pathlib import Path\n" + raw_code.replace(
                    "current_dir = Path(os.path.dirname(os.path.abspath(__file__)))",
                    "current_dir = Path(os.getcwd())"
                )

                exec_globals = globals().copy()
                exec(patched_code, exec_globals)

                # ── Progress callback: bridges exec'd engine → UI signal ──
                def progress_reporter(msg):
                    self.progress_signal.emit(msg)

                self.progress_signal.emit("🚀 Engine loaded — starting search...")
                self.progress_signal.emit("─" * 52)

                extraction_packet = exec_globals["extract_local_leads"](
                    search_query     = self.search_query,
                    allowed_ratings  = self.allowed_ratings,
                    target_city      = self.target_city,
                    progress_callback= progress_reporter   # ← NEW
                )
                self.finished_signal.emit(extraction_packet)
            else:
                self.error_signal.emit(f"Server error: {response.status_code}")

        except Exception as e:
            self.error_signal.emit(f"Fetch failed: {str(e)}")


# ── Deep investigator (email follow-up, future use) ───────────────
class DeepInvestigatorWorker(QThread):
    progress_signal = pyqtSignal(str)
    finished_signal = pyqtSignal(list)

    def __init__(self, pending_businesses):
        super().__init__()
        self.pending_businesses = pending_businesses

    def run(self):
        updated_data = []
        for business in self.pending_businesses:
            self.progress_signal.emit(f"Investigating: {business['name']}...")
            business['email'] = "Searching..."
            updated_data.append(business)
        self.finished_signal.emit(updated_data)


# ── Main UI ───────────────────────────────────────────────────────
class TimeVehicleUI(QWidget):
    def __init__(self):
        super().__init__()
        self.live_credits   = "Fetching Status..."
        self.scraper_worker = None
        self.init_ui()
        QTimer.singleShot(100, self.lazy_load_credits)

    def init_ui(self):
        self.setWindowTitle("Time Vehicle - 1.0")

        # Window icon
        if getattr(sys, 'frozen', False):
            base_dir = os.path.dirname(sys.executable)
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))
        icon_path = os.path.join(base_dir, "timevehicle.ico")
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))

        self.setMinimumSize(440, 600)
        self.resize(480, 720)
        self.setStyleSheet("background-color: #FFFFFF;")

        title_font   = QFont("Arial", 14, QFont.Weight.Bold)
        section_font = QFont("Arial", 11, QFont.Weight.Bold)
        label_font   = QFont("Arial", 10)

        outer_layout = QVBoxLayout(self)
        outer_layout.setContentsMargins(0, 0, 0, 0)
        outer_layout.setSpacing(0)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll_area.setStyleSheet("QScrollArea { border: none; background-color: #FFFFFF; }")

        scroll_content = QWidget()
        scroll_content.setStyleSheet("background-color: #FFFFFF;")
        master_layout = QVBoxLayout(scroll_content)
        master_layout.setContentsMargins(0, 0, 0, 0)
        master_layout.setSpacing(0)

        # ── Header bar ────────────────────────────────────────────
        header_widget = QWidget()
        header_widget.setStyleSheet("background-color: #002D4A; border: none;")
        header_layout = QHBoxLayout(header_widget)
        header_layout.setContentsMargins(25, 20, 25, 20)
        header_layout.setSpacing(15)
        self.logo_emblem = LogoWidget()
        header_layout.addWidget(self.logo_emblem)
        title = QLabel("TIME VEHICLE - 1.0")
        title.setFont(title_font)
        title.setStyleSheet("color: #FFFFFF; line-height: 1.2;")
        header_layout.addWidget(title)
        header_layout.addStretch()
        master_layout.addWidget(header_widget)

        # ── Main form area ────────────────────────────────────────
        sheet_widget = QWidget()
        sheet_widget.setStyleSheet("background-color: #FFFFFF;")
        sheet_layout = QVBoxLayout(sheet_widget)
        sheet_layout.setSpacing(12)
        sheet_layout.setContentsMargins(25, 15, 25, 20)

        # Credits bar
        credit_frame = QFrame()
        credit_frame.setStyleSheet(
            "background-color: #F8FAFC; border: 1px solid #E2E8F0; border-radius: 6px; padding: 4px;"
        )
        credit_layout = QHBoxLayout(credit_frame)
        credit_layout.setContentsMargins(12, 6, 12, 6)

        self.lbl_credits = QLabel(f"💳 Searches Remaining: {self.live_credits}", font=label_font)
        self.lbl_credits.setStyleSheet("color: #0F172A; font-weight: bold;")

        btn_recharge = QPushButton("🔌 Recharge")
        btn_recharge.setFont(QFont("Arial", 9, QFont.Weight.Bold))
        btn_recharge.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_recharge.setStyleSheet("""
            QPushButton { background-color: #10B981; color: white; border: none;
                          border-radius: 4px; padding: 5px 12px; }
            QPushButton:hover { background-color: #059669; }
        """)
        btn_recharge.clicked.connect(self.open_serpapi_billing)
        credit_layout.addWidget(self.lbl_credits)
        credit_layout.addStretch()
        credit_layout.addWidget(btn_recharge)
        sheet_layout.addWidget(credit_frame)

        # Input fields
        input_style = (
            "padding: 8px; border: 1px solid #B0C4DE; border-radius: 4px;"
            "background: #FFFFFF; color: #000000;"
        )

        lbl_prof = QLabel("Field of search / profession:", font=label_font)
        lbl_prof.setStyleSheet("color: #333333;")
        sheet_layout.addWidget(lbl_prof)
        self.input_profession = QLineEdit()
        self.input_profession.setPlaceholderText("e.g., Orthopedic doctor")
        self.input_profession.setStyleSheet(input_style)
        sheet_layout.addWidget(self.input_profession)

        lbl_local = QLabel("Locality / area name:", font=label_font)
        lbl_local.setStyleSheet("color: #333333;")
        sheet_layout.addWidget(lbl_local)
        self.input_locality = QLineEdit()
        self.input_locality.setPlaceholderText("e.g., Manhattan")
        self.input_locality.setStyleSheet(input_style)
        sheet_layout.addWidget(self.input_locality)

        geo_layout = QHBoxLayout()
        geo_layout.setSpacing(15)

        vbox_city = QVBoxLayout()
        lbl_city = QLabel("City:", font=label_font)
        lbl_city.setStyleSheet("color: #333333;")
        vbox_city.addWidget(lbl_city)
        self.input_city = QLineEdit()
        self.input_city.setPlaceholderText("e.g., New York")
        self.input_city.setStyleSheet(input_style)
        vbox_city.addWidget(self.input_city)

        vbox_state = QVBoxLayout()
        lbl_state = QLabel("State:", font=label_font)
        lbl_state.setStyleSheet("color: #333333;")
        vbox_state.addWidget(lbl_state)
        self.input_state = QLineEdit()
        self.input_state.setPlaceholderText("e.g., New York")
        self.input_state.setStyleSheet(input_style)
        vbox_state.addWidget(self.input_state)

        vbox_country = QVBoxLayout()
        lbl_country = QLabel("Country:", font=label_font)
        lbl_country.setStyleSheet("color: #333333;")
        vbox_country.addWidget(lbl_country)
        self.input_country = QLineEdit()
        self.input_country.setPlaceholderText("e.g., USA")
        self.input_country.setStyleSheet(input_style)
        vbox_country.addWidget(self.input_country)

        geo_layout.addLayout(vbox_city,    stretch=1)
        geo_layout.addLayout(vbox_state,   stretch=1)
        geo_layout.addLayout(vbox_country, stretch=1)
        sheet_layout.addLayout(geo_layout)

        self.add_separator(sheet_layout)

        # Rating checkboxes
        rating_label = QLabel("Google Rating Range")
        rating_label.setFont(section_font)
        rating_label.setStyleSheet("color: #002D4A;")
        sheet_layout.addWidget(rating_label)

        rating_layout = QHBoxLayout()
        ratings = ["5", "4", "3", "2", "1", "0", "ALL"]
        self.rating_checkboxes = {}
        checkbox_style = (
            "QCheckBox { color: #333333; padding: 4px; }"
            "QCheckBox::indicator { width: 16px; height: 16px; }"
        )
        for rate in ratings:
            cb = QCheckBox(rate)
            cb.setFont(label_font)
            cb.setStyleSheet(checkbox_style)
            if rate == "ALL":
                cb.setChecked(True)
                cb.toggled.connect(self.handle_all_ratings_toggle)
            else:
                cb.toggled.connect(self.handle_single_rating_toggle)
            rating_layout.addWidget(cb)
            self.rating_checkboxes[rate] = cb
        sheet_layout.addLayout(rating_layout)

        self.add_separator(sheet_layout)

        # Export format
        download_label = QLabel("Download to System:")
        download_label.setFont(section_font)
        download_label.setStyleSheet("color: #002D4A;")
        sheet_layout.addWidget(download_label)

        format_layout = QHBoxLayout()
        self.chk_excel = QCheckBox("Excel")
        self.chk_excel.setChecked(True)
        self.chk_excel.setFont(label_font)
        self.chk_excel.setStyleSheet(checkbox_style)
        format_layout.addWidget(self.chk_excel)
        sheet_layout.addLayout(format_layout)

        self.add_separator(sheet_layout)
        sheet_layout.addSpacing(5)

        # Submit button
        self.btn_submit = QPushButton("SUBMIT")
        self.btn_submit.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        self.btn_submit.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_submit.setStyleSheet("""
            QPushButton {
                background-color: #002D4A; color: white;
                padding: 12px; border: none; border-radius: 4px;
            }
            QPushButton:hover    { background-color: #004473; }
            QPushButton:pressed  { background-color: #001524; }
            QPushButton:disabled { background-color: #64748B; color: #CBD5E1; }
        """)
        self.btn_submit.clicked.connect(self.handle_submit_action)
        sheet_layout.addWidget(self.btn_submit)

        # ── LIVE PROGRESS LOG PANEL ───────────────────────────────
        # Visible only during an active search.
        self.progress_frame = QFrame()
        self.progress_frame.setStyleSheet(
            "QFrame { background-color: #F0F4F8; border: 1px solid #B0C4DE;"
            "border-radius: 6px; }"
        )
        progress_inner = QVBoxLayout(self.progress_frame)
        progress_inner.setContentsMargins(10, 8, 10, 8)
        progress_inner.setSpacing(5)

        log_header = QHBoxLayout()
        lbl_log_title = QLabel("📊  Live Progress Log")
        lbl_log_title.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        lbl_log_title.setStyleSheet("color: #002D4A; border: none; background: transparent;")
        self.lbl_lead_count = QLabel("Leads collected: 0")
        self.lbl_lead_count.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        self.lbl_lead_count.setStyleSheet("color: #059669; border: none; background: transparent;")
        log_header.addWidget(lbl_log_title)
        log_header.addStretch()
        log_header.addWidget(self.lbl_lead_count)
        progress_inner.addLayout(log_header)

        self.progress_log = QTextEdit()
        self.progress_log.setReadOnly(True)
        self.progress_log.setFont(QFont("Consolas", 9))
        self.progress_log.setFixedHeight(190)
        self.progress_log.setStyleSheet(
            "QTextEdit { background-color: #0F1E2A; color: #90EE90;"
            "border: none; border-radius: 4px; padding: 6px; }"
        )
        progress_inner.addWidget(self.progress_log)

        self.progress_frame.setVisible(False)   # hidden until search starts
        sheet_layout.addWidget(self.progress_frame)
        # ─────────────────────────────────────────────────────────

        sheet_layout.addSpacing(15)

        # Support footer
        support_frame = QFrame()
        support_frame.setStyleSheet(
            "background-color: #F8FAFC; border: 1px solid #E2E8F0; border-radius: 6px; padding: 12px;"
        )
        support_layout = QHBoxLayout(support_frame)
        lbl_support_text = QLabel(
            "💬 Need help or customized solutions? WhatsApp us at:\n"
            "👉 +91 77803 79259\n"
            "📧 Mail us at: support@timevehicle.com"
        )
        lbl_support_text.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        lbl_support_text.setStyleSheet("color: #475569; line-height: 1.4;")
        lbl_support_text.setAlignment(Qt.AlignmentFlag.AlignCenter)
        support_layout.addWidget(lbl_support_text)
        sheet_layout.addWidget(support_frame)

        master_layout.addWidget(sheet_widget)
        scroll_area.setWidget(scroll_content)
        outer_layout.addWidget(scroll_area)

    # ── Helpers ───────────────────────────────────────────────────
    def add_separator(self, layout):
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setFrameShadow(QFrame.Shadow.Sunken)
        line.setStyleSheet(
            "color: #E2E8F0; margin-top: 5px; margin-bottom: 5px;"
            "border: none; background-color: #E2E8F0; height: 1px;"
        )
        layout.addWidget(line)

    def handle_all_ratings_toggle(self, checked):
        if checked:
            for rate, cb in self.rating_checkboxes.items():
                if rate != "ALL":
                    cb.blockSignals(True)
                    cb.setChecked(False)
                    cb.blockSignals(False)

    def handle_single_rating_toggle(self, checked):
        if checked:
            self.rating_checkboxes["ALL"].blockSignals(True)
            self.rating_checkboxes["ALL"].setChecked(False)
            self.rating_checkboxes["ALL"].blockSignals(False)

    def open_serpapi_billing(self, event=None):
        webbrowser.open("https://serpapi.com/plan")
        self.refresh_credit_display()

    def lazy_load_credits(self):
        self.live_credits = fetch_live_serp_credits()
        self.lbl_credits.setText(f"💳 Searches Remaining: {self.live_credits}")

    def refresh_credit_display(self):
        self.lbl_credits.setText("💳 Searches Remaining: Fetching...")
        self.live_credits = fetch_live_serp_credits()
        self.lbl_credits.setText(f"💳 Searches Remaining: {self.live_credits}")

    # ── Passkey verification ──────────────────────────────────────
    def verify_security_passkey(self):
        """
        First run:       prompts for passkey → validates → saves locally.
        Every run after: silently re-validates the saved key in background.
        Revoked key:     clears cache, prompts for a new one.
        Offline:         saved key is trusted (prevents blocking clients
                         with a temporary internet drop mid-session).
        """
        saved = _load_saved_passkey()
        if saved:
            result = _validate_passkey_cloud(saved)
            if result is True:
                return True          # ← silent pass, client sees nothing
            elif result is None:
                return True          # ← offline but key was previously valid
            else:
                # Key revoked — clear cache, fall through to prompt
                _clear_saved_passkey()
                QMessageBox.warning(
                    self, "Access Key Expired",
                    "Your Time Vehicle access key has expired or been revoked.\n"
                    "Please enter a new passkey to continue."
                )

        # First run or revoked key — prompt
        passkey, ok = QInputDialog.getText(
            self, "Security Verification",
            "Please enter your Time Vehicle Activation Passkey:",
            QLineEdit.EchoMode.Normal
        )
        if not ok:
            return False
        passkey_clean = passkey.strip()
        if not passkey_clean:
            QMessageBox.warning(self, "Entry Empty", "Passkey cannot be empty.")
            return False

        result = _validate_passkey_cloud(passkey_clean)
        if result is True:
            _save_passkey(passkey_clean)   # ← save so client never sees this again
            return True
        elif result is None:
            QMessageBox.critical(
                self, "Network Failure",
                "Could not reach the validation server.\n"
                "Please check your internet connection and try again."
            )
            return False
        else:
            QMessageBox.critical(
                self, "Verification Failure",
                "The passkey you entered is not authorized or has expired."
            )
            return False

    # ── Submit handler ────────────────────────────────────────────
    def handle_submit_action(self):
        profession = self.input_profession.text().strip()
        locality   = self.input_locality.text().strip()
        city       = self.input_city.text().strip()
        state      = self.input_state.text().strip()
        country    = self.input_country.text().strip()

        if not profession or not city:
            QMessageBox.warning(
                self, "Missing Information",
                "Profession and City fields are mandatory!"
            )
            return

        selected_ratings = []
        if self.rating_checkboxes["ALL"].isChecked():
            selected_ratings = ["ALL"]
        else:
            for rate, cb in self.rating_checkboxes.items():
                if rate != "ALL" and cb.isChecked():
                    selected_ratings.append(rate)

        if not selected_ratings:
            QMessageBox.warning(
                self, "Missing Rating",
                "Please select at least one rating or 'ALL'."
            )
            return

        if not self.verify_security_passkey():
            return

        # ── Chromium install (works as both .py script and frozen .exe) ──
        # When frozen, sys.executable is the .exe itself — running
        # "TimeVehicle.exe -m playwright" would crash. Instead, we use
        # Playwright's own bundled driver executable directly.
        self.btn_submit.setText("⚙️ PREPARING BROWSER... PLEASE WAIT")
        self.btn_submit.repaint()
        try:
            if getattr(sys, 'frozen', False):
                from playwright._impl._driver import compute_driver_executable
                pw_driver = str(compute_driver_executable())
                subprocess.check_call([pw_driver, "install", "chromium"])
            else:
                subprocess.check_call([sys.executable, "-m", "playwright", "install", "chromium"])
        except Exception as e:
            QMessageBox.warning(
                self, "Setup Notice",
                "Browser setup failed. Please ensure internet is connected.\n"
                f"Detail: {str(e)}"
            )
            self.btn_submit.setText("SUBMIT")
            self.btn_submit.setEnabled(True)
            return

        # Build query strings
        search_components = [profession]
        if locality:
            search_components.append(locality)
        clean_search_query = " ".join(search_components)

        geo_components = [city]
        if state:   geo_components.append(state)
        if country: geo_components.append(country)
        target_location_context = ", ".join(geo_components)

        # Prepare progress panel
        self.progress_log.clear()
        self.lbl_lead_count.setText("Leads collected: 0")
        self.progress_frame.setVisible(True)

        self.btn_submit.setEnabled(False)
        self.btn_submit.setText("⏳ DOWNLOADING DATA... PLEASE WAIT")

        self.scraper_worker = ScraperWorker(
            clean_search_query, selected_ratings, target_location_context
        )
        self.scraper_worker.finished_signal.connect(self.on_scraping_finished)
        self.scraper_worker.error_signal.connect(self.on_scraping_error)
        self.scraper_worker.progress_signal.connect(self.on_progress_update)  # ← NEW
        self.scraper_worker.start()

    # ── Live progress slot ────────────────────────────────────────
    def on_progress_update(self, message):
        """Appends a log line and auto-scrolls. Updates lead counter."""
        self.progress_log.append(message)
        # Auto-scroll to bottom
        sb = self.progress_log.verticalScrollBar()
        sb.setValue(sb.maximum())

        # Parse lead count from lines like "🏢 [12] Business Name ★4.5"
        if message.startswith("🏢 ["):
            try:
                count = int(message.split("[")[1].split("]")[0])
                self.lbl_lead_count.setText(f"Leads collected: {count}")
            except Exception:
                pass

        # Final summary line
        if "DONE —" in message:
            try:
                total = int(message.split("DONE —")[1].split("qualified")[0].strip())
                self.lbl_lead_count.setText(f"✅ Total leads: {total}")
            except Exception:
                pass

    # ── Finished handler ──────────────────────────────────────────
    def on_scraping_finished(self, extraction_packet):
        try:
            extracted_data       = extraction_packet.get("data", [])
            active_columns_layout = extraction_packet.get("columns_layout", None)

            if not extracted_data:
                QMessageBox.warning(
                    self, "No Results",
                    "No businesses found matching your parameters.\n"
                    "Try broadening your search or changing the rating filter."
                )
                return

            saved_file_paths = []
            if self.chk_excel.isChecked():
                headers = {'User-Agent': 'Mozilla/5.0'}
                CLOUD_EXPORT_URL = (
                    "https://raw.githubusercontent.com/skengeneral/"
                    "TimeVehicle-basic/main/export_engine.py"
                )
                response = requests.get(CLOUD_EXPORT_URL, headers=headers, timeout=10)

                if response.status_code == 200:
                    export_code = response.text
                    export_code = export_code.replace("from openpyxl import Workbook", "")
                    export_code = export_code.replace(
                        "from openpyxl.styles import Font, PatternFill, Alignment, Border, Side", ""
                    )
                    export_code = export_code.replace("os.path.abspath(__file__)", "os.getcwd()")
                    export_code = export_code.replace("__file__", "os.path.abspath(os.getcwd())")

                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                    local_export_scope = {
                        "__builtins__": __builtins__,
                        "Workbook": Workbook,
                        "Font": Font,
                        "PatternFill": PatternFill,
                        "Alignment": Alignment,
                        "Border": Border,
                        "Side": Side,
                        "os": os,
                        "sys": sys,
                    }
                    exec(export_code, local_export_scope)
                    xl_path = local_export_scope["save_to_excel"](
                        extracted_data, active_columns_layout
                    )
                    saved_file_paths.append(f"• Excel File: {xl_path}")
                else:
                    raise Exception(f"Export server returned: {response.status_code}")

            self.refresh_credit_display()
            success_summary = (
                "✅  Leads Compiled Successfully!\n\n"
                f"Total Rows Gathered: {len(extracted_data)}\n\n"
                "Saved to:\n" + "\n".join(saved_file_paths)
            )
            QMessageBox.information(self, "Extraction Complete", success_summary)

        except Exception as e:
            QMessageBox.critical(
                self, "Export Failure",
                f"Failed to save your results: {str(e)}"
            )
        finally:
            self.btn_submit.setEnabled(True)
            self.btn_submit.setText("SUBMIT")

    def on_scraping_error(self, error_message):
        QMessageBox.critical(
            self, "Extraction Failure",
            f"An error stopped the search:\n{error_message}"
        )
        self.btn_submit.setEnabled(True)
        self.btn_submit.setText("SUBMIT")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    if sys.platform == "win32":
        app.setStyle('Fusion')
    window = TimeVehicleUI()
    window.show()
    sys.exit(app.exec())
