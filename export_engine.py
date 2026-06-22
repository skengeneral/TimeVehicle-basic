import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def save_to_excel(data, columns_layout=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Local Business Leads"
    
    # Enable explicit grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Professional Palette Design
    header_fill = PatternFill(start_color="002D4A", end_color="002D4A", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10, color="000000")
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border_side = Side(border_style="thin", color="E2E8F0")
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Structured standard headers mapping exactly to lead_card schema
    headers = [
        "Business Name", "Google Rating", "Complete Address", "Operating Hours Matrix",
        "Website Link", "Email ID", "Phone Number", "Facebook Handle",
        "Instagram Handle", "LinkedIn Handle", "Twitter/X Handle"
    ]
    
    ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = cell_border
    ws.row_dimensions[1].height = 28
    
    # Populate rows
    for row_idx, item in enumerate(data, 2):
        row_data = [item.get(h, "Not Provided") for h in headers]
        ws.append(row_data)
        ws.row_dimensions[row_idx].height = 22
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = cell_border
            # Format Google Rating column symmetrically
            if headers[col_idx-1] == "Google Rating":
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # Set column widths dynamically
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)
        
    # Generate output path next to the user's workspace executable
    if getattr(sys, 'frozen', False):
        base_path = os.path.dirname(sys.executable)
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))
        
    filename = os.path.join(base_path, "Time_Vehicle_Leads.xlsx")
    wb.save(filename)
    return filename
